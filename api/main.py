from fastapi import FastAPI, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
import io
from typing import List, Dict, Optional
from pydantic import BaseModel

app = FastAPI()

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class CellStyle(BaseModel):
    backgroundColor: Optional[str] = None
    color: Optional[str] = None
    fontWeight: Optional[str] = None
    fontStyle: Optional[str] = None
    textDecoration: Optional[str] = None

class CellInfo(BaseModel):
    content: str
    style: CellStyle
    address: str

class SheetData(BaseModel):
    name: str
    cells: List[CellInfo]

def rgb_to_hex(rgb_value) -> str:
    """Convert RGB value to hex color code."""
    if not rgb_value:
        return None
    
    # Handle openpyxl RGB object
    if hasattr(rgb_value, 'rgb'):
        rgb_str = rgb_value.rgb
        if not rgb_str:
            return None
        # If it's already a hex color (starts with FF or 00), return it
        if len(rgb_str) == 8:
            return f"#{rgb_str[2:]}"
        return f"#{rgb_str}"
    
    # Handle string input
    if isinstance(rgb_value, str):
        if rgb_value.startswith("RGB"):
            # Parse "RGB(r,g,b)" format
            rgb = rgb_value.strip("RGB()").split(",")
            r, g, b = map(int, rgb)
            return f"#{r:02x}{g:02x}{b:02x}"
        elif rgb_value.startswith("#"):
            # Already in hex format
            return rgb_value
        else:
            try:
                # Try to convert direct color code
                if len(rgb_value) == 8:  # ARGB format
                    return f"#{rgb_value[2:]}"
                return f"#{rgb_value}"
            except:
                return None
    
    return None

def get_cell_style(cell) -> CellStyle:
    """Extract style information from a cell."""
    style = CellStyle()
    
    try:
        if cell.font:
            style.fontWeight = "bold" if cell.font.bold else None
            style.fontStyle = "italic" if cell.font.italic else None
            style.textDecoration = "underline" if cell.font.underline else None
            if cell.font.color:
                try:
                    if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                        rgb_str = cell.font.color.rgb
                        if isinstance(rgb_str, str):
                            style.color = f"#{rgb_str}" if not rgb_str.startswith('#') else rgb_str
                        elif hasattr(rgb_str, 'hex'):
                            style.color = f"#{rgb_str.hex()}"
                except Exception as e:
                    print(f"Error processing font color: {e}")
        
        if cell.fill and cell.fill.patternType == "solid" and cell.fill.fgColor:
            try:
                if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                    rgb_str = cell.fill.fgColor.rgb
                    if isinstance(rgb_str, str):
                        style.backgroundColor = f"#{rgb_str}" if not rgb_str.startswith('#') else rgb_str
                    elif hasattr(rgb_str, 'hex'):
                        style.backgroundColor = f"#{rgb_str.hex()}"
            except Exception as e:
                print(f"Error processing background color: {e}")
    except Exception as e:
        print(f"Error processing cell style: {str(e)}")
    
    return style

@app.post("/api/analyze-sheet")
async def analyze_sheet(
    file: UploadFile = File(...),
    sheet_name: str = Form(None)
) -> SheetData:
    # Read the Excel file
    contents = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    
    print(f"Received sheet_name: {sheet_name}")
    print(f"Available sheets: {workbook.sheetnames}")
    
    # Use first sheet if sheet_name not provided or not found
    if not sheet_name or sheet_name not in workbook.sheetnames:
        sheet_name = workbook.sheetnames[0]
        print(f"Using default sheet: {sheet_name}")
    
    worksheet = workbook[sheet_name]
    cells: List[CellInfo] = []
    
    # Get the actual used range of the worksheet
    min_row = worksheet.min_row
    max_row = worksheet.max_row
    min_col = worksheet.min_column
    max_col = worksheet.max_column
    
    print(f"Processing sheet: {sheet_name}")
    print(f"Range: rows {min_row}-{max_row}, cols {min_col}-{max_col}")
    
    # Process each cell in the used range
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            
            # Skip empty cells
            if cell.value is None:
                continue
            
            style = get_cell_style(cell)
            
            cell_info = CellInfo(
                content=str(cell.value),
                style=style,
                address=cell.coordinate
            )
            cells.append(cell_info)
    
    return SheetData(name=sheet_name, cells=cells)

@app.post("/api/sheets")
async def get_sheets(file: UploadFile = File(...)) -> List[str]:
    contents = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=False)
    return workbook.sheetnames 